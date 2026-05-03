from io import BytesIO
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string
from models.excel import ExcelUpdateItem

def apply_excel_highlights(content: bytes, update_data: List[ExcelUpdateItem]) -> BytesIO:
    # 1. 파일 로드
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    # 2. 엑셀 스타일 정의
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_row_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")   # F2F2F2
    font = Font(color="000000", bold=True, size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 3. 변경된 행 전체 배경 적용
    modified_rows = set([item.row for item in update_data])
    max_col = ws.max_column

    for r in modified_rows:
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = light_row_fill

    # 4. 개별 셀 데이터 반영 및 스타일링
    for item in update_data:
        r, c_letter, val = item.row, item.column, item.value
        
        # 엑셀 열 문자("AX")를 숫자(50)로 변환 (1부터 시작)
        c_idx = column_index_from_string(c_letter)

        # [openpyxl] 셀 업데이트 및 스타일 적용
        cell = ws.cell(row=r, column=c_idx)
        cell.value = val

        cell.fill = fill
        cell.font = font
        cell.border = thin_border

    # 3. 수정된 엑셀 파일을 메모리에 저장하여 반환
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output