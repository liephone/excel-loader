# Excel 스타일 적용 예제
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill


wb = load_workbook("C:\\Users\\lieph\\repos\\ledger-data-generator\\가데이터작성.xlsx")
ws = wb.active

font_style = Font(name='맑은 고딕', size=10, bold=True, color='000000')
fill_style = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 
border_style = Border(top=Side(border_style="thin", color="000000"),
                    #   left=Side(border_style="thin", color="000000"),
                    #   right=Side(border_style="thin", color="000000"),
                      bottom=Side(border_style="thin", color="000000"))
alignment_style = Alignment(horizontal='center', vertical='center')

for row in ws['A3':'AX3']:
    for cell in row:
        cell.font = font_style
        # cell.fill = fill_style
        cell.border = border_style
        # cell.alignment = alignment_style

# cell = ws['A2':'AX2']
# cell.font = font_style
# cell.border = border_style
# cell.alignment = alignment_style

# # 열 너비 조절 (A열)
# ws.column_dimensions['A'].width = 30

# # 행 높이 조절 (1행)
# ws.row_dimensions[1].height = 40

wb.save("C:\\Users\\lieph\\repos\\ledger-data-generator\\가데이터작성_styled.xlsx")
print("Excel 스타일 적용 완료!")

# wb = Workbook()
# ws = wb.active
# cell = ws['A1']
# cell.value = 'Hello, World!'

# # 1. 폰트 설정 (굵게, 빨간색)
# cell.font = Font(name='맑은 고딕', size=14, bold=True, color='FF0000')

# # 2. 정렬 설정 (가운데)
# cell.alignment = Alignment(horizontal='center', vertical='center')

# # 3. 테두리 설정 (얇은 테두리)
# thin = Side(border_style="thin", color="000000")
# cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

# # 4. 배경색 설정 (노란색)
# cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# wb.save("styled_excel.xlsx")