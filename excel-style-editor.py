import streamlit as st
import pandas as pd
import json
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string


# 1. 사이드바 설정 (하이라이팅 옵션)
st.sidebar.header("🎨 스타일 설정")
bg_color = st.sidebar.color_picker("배경색 선택", "#FFFF00")  # 기본 노란색
font_color = st.sidebar.color_picker("글자색 선택", "#000000")
is_bold = st.sidebar.checkbox("굵게(Bold)", value=True)
font_size = st.sidebar.number_input("글자 크기", value=11, help="글자 크기는 다운로드 파일에서만 확인 가능합니다.")
use_border = st.sidebar.toggle("테두리 적용", value=True, help="테두리 스타일은 다운로드 파일에서만 확인 가능합니다.")

# 2. 파일 업로드
st.title("📊 Excel Data Updater & Highlighter")
excel_file = st.file_uploader("엑셀 파일을 업로드하세요", type=["xlsx"])
json_file = st.file_uploader("수정 정보 JSON 파일을 업로드하세요", type=["json"])

# 샘플 JSON 구조 안내 (확장성)
with st.expander("JSON 파일 예시 형식"):
    st.code("""
[
    {"row": 2, "column": "R", "value": "수정된 값1"},
    {"row": 5, "column": "X", "value": "수정된 값2"}
]
    """)

if excel_file and json_file:
    # 데이터 로드
    df = pd.read_excel(excel_file)
    update_data = json.load(json_file)
    wb = load_workbook(excel_file)
    ws = wb.active

    # 컬럼 매핑 (이름 -> 엑셀 열 번호)
    # col_map = {col: i + 1 for i, col in enumerate(df.columns)}

    # 스타일 정의
    fill = PatternFill(start_color=bg_color[1:], end_color=bg_color[1:], fill_type="solid")
    light_row_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")   # F2F2F2
    font = Font(color=font_color[1:], bold=is_bold, size=font_size)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 수정된 셀 좌표 추적 (Pandas Styler용)
    updated_cells = []

    # 변경된 행에 연한 회색 배경 적용 (light_row_fill)
    modified_rows = set([item['row'] for item in update_data])
    max_col = ws.max_column

    for r in modified_rows:
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = light_row_fill

    # 데이터 반영 및 스타일링 (openpyxl)
    for item in update_data:
        # r, c_name, val = item['row'], item['column'], item['value']
        r, c_letter, val = item['row'], item['column'], item['value']
        
        try:
            # 엑셀 열 문자("AX")를 숫자(50)로 변환 (1부터 시작)
            c_idx = column_index_from_string(c_letter)

            # [openpyxl] 셀 업데이트 및 스타일 적용
            cell = ws.cell(row=r, column=c_idx)
            cell.value = val

            cell.fill = fill
            cell.font = font
            if use_border:
                cell.border = thin_border

            # [Pandas] 데이터프레임(df)에도 직접 반영 (화면 출력용)
            df_row_idx = r - 2      # 엑셀의 1행이 헤더인 경우, 데이터는 2행부터 시작
            df_col_idx = c_idx - 1  # 엑셀 열 번호는 1부터 시작, DataFrame은 0부터 시작

            if 0 <= df_row_idx < len(df) and 0 <= df_col_idx < len(df.columns):
                # iat[행 인덱스, 열 인덱스]로 값 업데이트
                df.iat[df_row_idx, df_col_idx] = val  # DataFrame 열 인덱스는 0부터 시작
                updated_cells.append((df_row_idx, df.columns[df_col_idx]))  # (행 인덱스, 열 이름)

        except ValueError:
            st.error(f"잘못된 엑셀 컬럼 문자입니다: {c_letter}")

        # if c_name in col_map:
        #     c_idx = col_map[c_name]
        #     cell = ws.cell(row=r, column=c_idx)
        #     cell.value = val

        #     cell.fill = fill
        #     cell.font = font
        #     cell.border = thin_border

        #     # [B] 데이터프레임(df)에도 직접 반영 (화면 출력용)
        #     # 엑셀의 row 2가 데이터의 0번째 인덱스라고 가정 (r-2)
        #     df_row_idx = r - 2 
        #     df.at[df_row_idx, c_name] = val
        
        #     # Pandas 인덱스로 변환 (엑셀 1번행이 헤더인 경우 등 고려 필요)
        #     # 여기서는 단순 예시를 위해 엑셀 좌표 그대로 기록
        #     updated_cells.append((df_row_idx, c_name))

    def highlight_modified(data):
        # 전체를 빈 문자열로 채운 데이터프레임 생성
        style_df = pd.DataFrame('', index=data.index, columns=data.columns)

        # 기록된 좌표에만 스타일 문자열 삽입
        for r_idx, c_name in updated_cells:
            style_str = f'background-color: {bg_color}; color: {font_color};'
            if is_bold: style_str += ' font-weight: bold;'
            style_str += f' font-size: {font_size}pt;'

            # 해당 좌표에 스타일 적용
            if r_idx in style_df.index and c_name in style_df.columns:
                style_df.at[r_idx, c_name] = style_str

        return style_df

    st.subheader("업데이트 결과 미리보기")
    st.dataframe(df.style.apply(highlight_modified, axis=None))

    # 5. 결과 파일 다운로드
    output = BytesIO()
    wb.save(output)
    processed_data = output.getvalue()

    st.download_button(
        label="📥 수정된 엑셀 다운로드",
        data=processed_data,
        file_name="updated_excel.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.success("반영이 완료되었습니다!")
